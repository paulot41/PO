"""
Permite importar e exportar Tarefas em ficheiros Excel (.xlsx).
"""
import openpyxl
from datetime import datetime
from tarefa import Tarefa

class ExcelManager:
    """
    Classe para lidar com importação/exportação Excel das tarefas.
    """
    def importar_tarefas(self, ficheiro_excel):
        wb = openpyxl.load_workbook(ficheiro_excel)
        sheet = wb.active
        tarefas = []
        # Espera-se cabeçalho: ID, Nome, Email, Telefone, Morada, Prioridade, Descricao, DataCriacao, Escalonada
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            (id_tarefa,
             nome,
             email,
             telefone,
             morada,
             prioridade,
             descricao,
             data_criacao_str,
             escalonada_flag) = row

            if not nome:
                # Se nem o nome está preenchido, ignoramos
                continue

            if data_criacao_str:
                try:
                    data_criacao = datetime.fromisoformat(data_criacao_str)
                except ValueError:
                    data_criacao = datetime.now()
            else:
                data_criacao = datetime.now()

            escalonada = (escalonada_flag == 1) if escalonada_flag else False

            if prioridade not in [1, 2, 3]:
                prioridade = 2

            tarefa = Tarefa(
                id=id_tarefa if id_tarefa else "",
                nome=nome,
                email=email if email else '',
                telefone=telefone if telefone else '',
                morada=morada if morada else '',
                prioridade=prioridade,
                descricao=descricao if descricao else '',
                data_criacao=data_criacao,
                escalonada=escalonada
            )
            tarefas.append(tarefa)
        return tarefas

    def exportar_tarefas(self, ficheiro_excel, tarefas):
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Cabeçalho
        sheet.append(["ID", "Nome", "Email", "Telefone", "Morada", "Prioridade", "Descricao", "DataCriacao", "Escalonada"])

        for tarefa in tarefas:
            sheet.append([
                tarefa.id,
                tarefa.nome,
                tarefa.email,
                tarefa.telefone,
                tarefa.morada,
                tarefa.prioridade,
                tarefa.descricao,
                tarefa.data_criacao.isoformat(),
                1 if tarefa.escalonada else 0
            ])
        wb.save(ficheiro_excel)
